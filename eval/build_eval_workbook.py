#!/usr/bin/env python3
"""Regenerate eval/MultiSourceAgent_Eval.xlsx — the BLANK 12-question scoring
template (Comment/Score columns stay empty; .filled.xlsx is the scored copy).

Final 12 (renumbered Q1-Q12). Gold answers verified 2026-06-17 against the
conformed c_* tables via the lh_supply_demo SQL endpoint. Re-run anytime:
    python eval/build_eval_workbook.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "MultiSourceAgent_Eval.xlsx")

# (#, Tier, Question, What it tests, Gold answer)
ROWS = [
    ("Q1", 1, "What was our revenue in Q1 2026 by channel?",
     "Both answer; baseline for token comparison",
     "Online $7.030M / Retail $7.156M / Wholesale $6.911M; total $21.097M (ERP, Jan-Mar 2026)"),
    ("Q2", 1, "What was our stockout rate in May 2026?",
     "Definition drift (share of SKUs vs share of rows)",
     "1.52% of SKU-location-day snapshots (ERP daily snapshots, May 2026)"),
    ("Q3", 2, "What were total company sales in May 2026, including Lakeside and the marketplace?",
     "Cross-source conformance + gross/net rule",
     "$10.054M = ERP $8.382M + marketplace NET $0.622M + Lakeside $1.051M"),
    ("Q4", 2, "Which products generate the most helpdesk tickets per 1,000 units sold?",
     "Free-text conformance + ratio convention",
     "SKU-0045 Titan Solenoid Valve 142.7/1K (2,041 tickets); SKU-0014 Titan Rocker Switch 115.9/1K (1,967); "
     "next SKU-0085 at 40.5. Convention: SKUs with >=1,000 units; 88% of ticket refs resolve"),
    ("Q5", 2, "What is the service parts attach rate for Hydraulics products?",
     "ITM code conformance",
     "8.1 parts per 100 units sold (16,156 parts / 199,192 Hydraulics units)"),
    ("Q6", 3, "Which A-class products have open complaints AND are at stockout risk in the merged network?",
     "The flagship multi-hop question",
     "SKU-0014 Titan Rocker Switch 20A and SKU-0045 Titan Solenoid Valve 1in "
     "(open complaint tickets; on-hand <= safety stock at 2026-05-31; near-zero Lakeside counts)"),
    ("Q7", 3, "Compare Lakeside store sell-through vs our DCs since the acquisition.",
     "Mixed-grain inventory + acquisition date",
     "Feb-May 2026: Lakeside 2.85x vs DCs 1.70x. Lakeside turns ~1.7x faster"),
    ("Q8", 3, "What were Lakeside sales on 03/02/2026?",
     "Date-format trap (DD/MM/YYYY)",
     "$36,836 (Feb 3, 2026 - dates are DD/MM/YYYY). An MM/DD misparse gives Mar 2 = $44,972"),
    ("Q9", 2, "How much revenue has Lakeside generated since the acquisition?",
     "Period convention from Prep-for-AI",
     "$4.324M (Feb 1 - May 31, 2026; ~9% from Lakeside store-brand items)"),
    ("Q10", 3, "How much of our total revenue can we actually tie to a known product?",
     "Governance / honest gap surfacing",
     "99.5% attributable; $0.553M (0.48% of $115.86M net) is unattributable - ALL of it stale MegaMart "
     "'UNMAPPED' listings (7.5% of marketplace net). Modeled reports the gap; raw assumes 100%"),
    ("Q11", 3, "Which product category has the worst after-sale reliability?",
     "Dual conformance (free-text + ITM) + category rollup + confidence note",
     "Hydraulics, 88.5 after-sale issues per 1K units (1,420 complaints + 16,156 service parts / 198,582 units); "
     "next Electrical 27.9. Confidence note: an 'Unknown' 2-product bucket scores 119.7/1K on only 3,811 units - exclude it"),
    ("Q12", 3, "Which products are most overstocked relative to their category's sell-through?",
     "Nested measure vs computed category benchmark, on mixed-grain inventory",
     "Worst overstock vs ~6x category avg: SKU-0029 Atlas Drop-In Anchor M8 (0.47x) & SKU-0039 Sterling Sleeve "
     "Anchor M12 (0.50x, Fasteners), SKU-0081 Atlas O-Ring Kit (0.50x, Hydraulics). No A-class sells below its "
     "category average (0/24); underperformers are B (25/31) and C (43/58) class"),
]

CONVENTIONS = [
    ("Convention", "Value"),
    ("Data window", "2025-06-01 to 2026-05-31"),
    ("Acquisition date (Lakeside)", "2026-02-01"),
    ("'Total / company sales'", "ERP + marketplace NET + Lakeside"),
    ("Marketplace revenue", "NET of fees (payoutAmount) unless 'gross' is asked"),
    ("Lakeside dates", "TEXT DD/MM/YYYY in raw exports"),
    ("Stockout rate", "share of SKU-location-day snapshots with is_stockout = 1 (grain = daily)"),
    ("Stockout risk", "on-hand <= safety stock at latest snapshot (ERP) or near-zero latest weekly count (Lakeside)"),
    ("Complaint", "helpdesk category = 'Product complaint' (is_complaint = TRUE)"),
    ("Tickets per 1K units", "resolved product refs only (88%); SKUs with >= 1,000 units sold"),
    ("Service parts attach rate", "service parts used / units sold x 100 (per 100 units); ITM-code conformance"),
    ("Marketplace fee rate", "14% of gross; net (payoutAmount) = gross - fees"),
    ("Product ABC class", "ERP product-master attribute (A/B/C by sales importance); Lakeside-only and special SKUs default to C"),
    ("Sell-through ratio", "units sold / average on-hand units over the period"),
    ("Attributable revenue", "net revenue whose sku is a real product (not 'UNMAPPED' / 'UNRESOLVED'); ~99.5% attributable, gap = stale MegaMart UNMAPPED listings"),
    ("After-sale issues per 1K", "(product complaints + service parts used) / units sold x 1000, by category"),
    ("Category avg sell-through", "mean of per-product sell-through within a category; overstock = product < category avg"),
    ("Token/CU measurement", "Capacity Metrics app (capacity admin on fabricassesmentcoe) or Foundry/SDK"),
]

# ---- styling ----
HDR_FILL = PatternFill("solid", fgColor="2E5496")
SUB_FILL = PatternFill("solid", fgColor="8EAADB")
NOTE_FILL = PatternFill("solid", fgColor="D9E1F2")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, fill=HDR_FILL):
    cell.fill = fill
    cell.font = WHITE_BOLD
    cell.alignment = CENTER
    cell.border = BORDER


wb = Workbook()

# ============================================================ Eval sheet
ws = wb.active
ws.title = "Eval"
hdr = ["#", "Tier", "Question", "What it tests", "Gold answer",
       "Raw_Plus", None, "Modeled", None]
for c, val in enumerate(hdr, 1):
    cell = ws.cell(row=1, column=c, value=val)
    style_header(cell)
# sub-header row 2
sub = [None, None, None, None, None, "Comment", "Score", "Comment", "Score"]
for c, val in enumerate(sub, 1):
    cell = ws.cell(row=2, column=c, value=val)
    style_header(cell, SUB_FILL)
# merges for the two-row header
for rng in ("A1:A2", "B1:B2", "C1:C2", "D1:D2", "E1:E2", "F1:G1", "H1:I1"):
    ws.merge_cells(rng)

r = 3
for qid, tier, q, tests, gold in ROWS:
    ws.cell(row=r, column=1, value=qid).font = BOLD
    ws.cell(row=r, column=2, value=tier).alignment = CENTER
    ws.cell(row=r, column=3, value=q)
    ws.cell(row=r, column=4, value=tests)
    ws.cell(row=r, column=5, value=gold)
    for c in range(1, 10):
        cc = ws.cell(row=r, column=c)
        cc.border = BORDER
        if c in (3, 4, 5, 6, 8):
            cc.alignment = WRAP_TOP
        elif c in (7, 9):
            cc.alignment = CENTER
    r += 1

note_row = r + 1
ws.cell(row=note_row, column=1,
        value="Scoring: 2 = correct, 1 = partially correct, 0 = wrong, -1 = hallucinated (confident + false)").font = BOLD
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)
ws.cell(row=note_row, column=1).fill = NOTE_FILL
agents_row = note_row + 1
ws.cell(row=agents_row, column=1,
        value="Agents: Raw_Plus = SupplyAgent_Raw_Plus (lh_supply_demo SQL endpoint + heavy agent instructions) | "
              "Modeled = SupplyAgent_Modeled (MultiSource_Modeled star + Prep-for-AI)").font = BOLD
ws.merge_cells(start_row=agents_row, start_column=1, end_row=agents_row, end_column=9)
ws.cell(row=agents_row, column=1).fill = NOTE_FILL

widths = {"A": 5, "B": 5, "C": 40, "D": 32, "E": 70, "F": 34, "G": 8, "H": 34, "I": 8}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A3"

# ============================================================ Summary sheet
# The score/percent/count cells are live formulas over the Eval score columns,
# so the totals recalc as you fill in scores (Excel evaluates them on open;
# openpyxl writes the formula text but does not compute it).
ws2 = wb.create_sheet("Summary")
shdr = ["Agent", "Data source", "Total score", "Max", "% of max",
        "Correct (2)", "Partial (1)", "Wrong (0)", "Hallucinated (-1)", "Answered"]
for c, val in enumerate(shdr, 1):
    cell = ws2.cell(row=1, column=c, value=val)
    cell.fill = PatternFill("solid", fgColor="1F3864")
    cell.font = WHITE_BOLD
    cell.alignment = CENTER
    cell.border = BORDER
# Score columns on the Eval sheet: Raw_Plus = G, Modeled = I; questions in rows 3..(2+len(ROWS))
FIRST_Q_ROW = 3
LAST_Q_ROW = 2 + len(ROWS)
# (Agent, Data source, Eval score column letter)
srows = [
    ("SupplyAgent_Raw_Plus", "lh_supply_demo SQL endpoint + heavy agent instructions", "G"),
    ("SupplyAgent_Modeled", "MultiSource_Modeled star + Prep-for-AI", "I"),
]
for i, (agent, source, col) in enumerate(srows, 2):
    rng = f"Eval!{col}{FIRST_Q_ROW}:{col}{LAST_Q_ROW}"
    values = [
        agent,                                  # A Agent
        source,                                 # B Data source
        f"=SUM({rng})",                         # C Total score
        24,                                     # D Max
        f"=IF(D{i}=0,\"\",C{i}/D{i})",          # E % of max
        f"=COUNTIF({rng},2)",                   # F Correct (2)
        f"=COUNTIF({rng},1)",                   # G Partial (1)
        f"=COUNTIF({rng},0)",                   # H Wrong (0)
        f"=COUNTIF({rng},-1)",                  # I Hallucinated (-1)
        f"=COUNT({rng})",                       # J Answered
    ]
    for c, val in enumerate(values, 1):
        cc = ws2.cell(row=i, column=c, value=val)
        cc.border = BORDER
        if c == 2:
            cc.alignment = WRAP_TOP
        elif c > 2:
            cc.alignment = CENTER
        if c == 5:  # % of max
            cc.number_format = "0.0%"
swidths = {"A": 22, "B": 50, "C": 11, "D": 6, "E": 9, "F": 11, "G": 10, "H": 10, "I": 16, "J": 10}
for col, w in swidths.items():
    ws2.column_dimensions[col].width = w

# ============================================================ Conventions sheet
ws3 = wb.create_sheet("Conventions")
for i, (k, v) in enumerate(CONVENTIONS, 1):
    kc = ws3.cell(row=i, column=1, value=k)
    vc = ws3.cell(row=i, column=2, value=v)
    vc.alignment = WRAP_TOP
    if i == 1:
        kc.fill = PatternFill("solid", fgColor="1F4E79")
        vc.fill = PatternFill("solid", fgColor="1F4E79")
        kc.font = WHITE_BOLD
        vc.font = WHITE_BOLD
    else:
        kc.font = BOLD
ws3.column_dimensions["A"].width = 26
ws3.column_dimensions["B"].width = 82

wb.save(OUT)
print("wrote", OUT, "| questions:", len(ROWS))
